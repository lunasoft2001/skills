#!/usr/bin/env python3
"""
bundle_manager.py — Presentation Bundle Manager
Scans a deliverables folder, generates index.xlsx and manifest.json.

Usage:
    python3 bundle_manager.py --slug my-slug
    python3 bundle_manager.py --slug my-slug --path /deliverables/my-slug \
                               --title "My Presentation" --author "Jane Smith"

Requirements:
    pip install openpyxl   (for index.xlsx — manifest.json has no dependencies)
"""

import argparse
import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

# ─── Core files expected in a complete bundle ─────────────────────────────────

CORE_FILES = {
    "storyboard.docx":    ("storyboard",    "Document"),
    "storyboard.json":    ("storyboard_json", "JSON"),
    "deck.pptx":          ("presentation",  "Presentation"),
    "speaker-notes.docx": ("speaker_notes", "Document"),
}

# ─── Helpers ──────────────────────────────────────────────────────────────────

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def file_info(path: Path) -> dict:
    stat = path.stat()
    return {
        "size_bytes": stat.st_size,
        "size_kb":    round(stat.st_size / 1024, 1),
        "created":    datetime.fromtimestamp(stat.st_ctime, tz=timezone.utc).isoformat(),
    }


def scan_folder(folder: Path) -> list[dict]:
    """Return info for every file in the folder (excluding index/manifest themselves)."""
    skip = {"index.xlsx", "manifest.json"}
    entries = []
    for f in sorted(folder.iterdir()):
        if f.is_file() and f.name not in skip:
            info = file_info(f)
            entries.append({
                "file":       f.name,
                "path":       f,
                "size_bytes": info["size_bytes"],
                "size_kb":    info["size_kb"],
                "created":    info["created"],
            })
    return entries

# ─── manifest.json ────────────────────────────────────────────────────────────

def build_manifest(slug: str, folder: Path, entries: list[dict],
                   project_title: str, author: str) -> dict:
    deliverables = []
    warnings = []
    all_present = True

    for fname, (ftype, _) in CORE_FILES.items():
        match = next((e for e in entries if e["file"] == fname), None)
        if match:
            deliverables.append({
                "file":       fname,
                "type":       ftype,
                "size_bytes": match["size_bytes"],
                "sha256":     sha256_file(match["path"]),
                "status":     "present",
            })
        else:
            all_present = False
            warnings.append(f"Missing core file: {fname}")
            deliverables.append({
                "file":   fname,
                "type":   ftype,
                "status": "missing",
            })

    # Extra files not in core list
    core_names = set(CORE_FILES.keys())
    for entry in entries:
        if entry["file"] not in core_names:
            deliverables.append({
                "file":       entry["file"],
                "type":       "other",
                "size_bytes": entry["size_bytes"],
                "sha256":     sha256_file(entry["path"]),
                "status":     "present",
            })

    return {
        "schema_version": "1.0",
        "project": {
            "title":        project_title or slug,
            "slug":         slug,
            "author":       author or "",
            "generated_at": datetime.now(tz=timezone.utc).isoformat(),
            "generated_by": "presentation-bundle-manager",
        },
        "deliverables": deliverables,
        "validation": {
            "all_core_files_present": all_present,
            "warnings":               warnings,
        },
    }

# ─── index.xlsx ───────────────────────────────────────────────────────────────

def build_index_xlsx(folder: Path, entries: list[dict], manifest: dict):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("WARNING: openpyxl not installed — index.xlsx skipped.")
        print("Install it with:  pip install openpyxl")
        return False

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(color="FFFFFF", bold=True)

    total_size_kb = sum(e["size_kb"] for e in entries)
    proj = manifest["project"]

    summary_rows = [
        ("Field", "Value"),
        ("Project",        proj["title"]),
        ("Slug",           proj["slug"]),
        ("Author",         proj["author"] or "—"),
        ("Generated",      proj["generated_at"]),
        ("Total files",    len(entries)),
        ("Total size (KB)", total_size_kb),
        ("All core files present", str(manifest["validation"]["all_core_files_present"])),
    ]

    for r_idx, (label, value) in enumerate(summary_rows, start=1):
        ws_sum.cell(r_idx, 1, label)
        ws_sum.cell(r_idx, 2, value)
        if r_idx == 1:
            for col in (1, 2):
                cell = ws_sum.cell(r_idx, col)
                cell.fill = header_fill
                cell.font = header_font

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 50

    # ── Sheet 2: Files ────────────────────────────────────────────────────────
    ws_files = wb.create_sheet("Files")
    col_headers = ["File", "Type", "Size (KB)", "Created", "Status", "Notes"]

    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws_files.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, d in enumerate(manifest["deliverables"], start=2):
        _, label_type = CORE_FILES.get(d["file"], ("other", "Other"))
        status_str = "✅ Present" if d["status"] == "present" else "❌ Missing"
        notes = ""

        ws_files.cell(row_idx, 1, d["file"])
        ws_files.cell(row_idx, 2, d.get("type", "other").replace("_", " ").title())
        ws_files.cell(row_idx, 3, round(d.get("size_bytes", 0) / 1024, 1))
        # created is only available in entries
        entry = next((e for e in entries if e["file"] == d["file"]), None)
        ws_files.cell(row_idx, 4, entry["created"] if entry else "—")
        ws_files.cell(row_idx, 5, status_str)
        ws_files.cell(row_idx, 6, notes)

    col_widths = [30, 18, 12, 30, 14, 20]
    for i, w in enumerate(col_widths, start=1):
        ws_files.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = folder / "index.xlsx"
    wb.save(str(output))
    print(f"✅ Saved index.xlsx → {output}")
    return True

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Bundle presentation deliverables")
    parser.add_argument("--slug",   required=True, help="Project slug (used as folder name)")
    parser.add_argument("--path",   help="Override deliverables path (default: /deliverables/<slug>)")
    parser.add_argument("--title",  default="", help="Human-readable project title")
    parser.add_argument("--author", default="", help="Author name")
    args = parser.parse_args()

    folder = Path(args.path) if args.path else Path(f"/deliverables/{args.slug}")

    if not folder.exists():
        print(f"ERROR: Deliverables folder not found: {folder}")
        sys.exit(1)

    entries = scan_folder(folder)

    if not entries:
        print(f"ERROR: No files found in {folder}")
        sys.exit(1)

    print(f"📁 Found {len(entries)} file(s) in {folder}")
    for e in entries:
        print(f"   {e['file']}  ({e['size_kb']} KB)")

    manifest = build_manifest(args.slug, folder, entries, args.title, args.author)

    # Write manifest.json (initial pass, without self-reference)
    manifest_path = folder / "manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    # Add manifest.json to entries and deliverables so it appears in index.xlsx
    # and in the manifest itself (re-write to include self-reference)
    manifest_info = file_info(manifest_path)
    entries.append({
        "file":       "manifest.json",
        "path":       manifest_path,
        "size_bytes": manifest_info["size_bytes"],
        "size_kb":    manifest_info["size_kb"],
        "created":    manifest_info["created"],
    })
    manifest["deliverables"].append({
        "file":       "manifest.json",
        "type":       "manifest",
        "size_bytes": manifest_info["size_bytes"],
        "sha256":     sha256_file(manifest_path),
        "status":     "present",
    })

    # Re-write manifest.json so the file on disk includes itself in deliverables
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f"✅ Saved manifest.json → {manifest_path}")

    if manifest["validation"]["warnings"]:
        for w in manifest["validation"]["warnings"]:
            print(f"⚠️  {w}")

    # Write index.xlsx
    build_index_xlsx(folder, entries, manifest)

    print("\n📦 Bundle complete.")
    if manifest["validation"]["all_core_files_present"]:
        print("   All core files present — bundle is complete.")
    else:
        print("   Some core files are missing — see warnings above.")


if __name__ == "__main__":
    main()
